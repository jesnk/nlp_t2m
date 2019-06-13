import openpyxl
import os
import numpy as np

#저장 할 파일 이름

def print_and_save_result_to_excel(testResult) :
    file_name = 'result.xlsx'

    if os.path.isfile(file_name):
    #file 있을시 load
        wb = openpyxl.load_workbook(file_name)
    # 데이터를 넣을 시트를 활성화합니다.
        sheet1 = wb.active
    else:
    #file 없을시 생성
        wb = openpyxl.Workbook()
    # 데이터를 넣을 시트를 활성화합니다.
        sheet1 = wb.active
    #시트의 이름을 정합니다.
        sheet1.title = 'resultSheet'


    # 넣을 list -> result_list
    result_list = testResult
    first_row = ["#########", "TESTIFY"]
    print(str(first_row))
    sheet1.append(first_row)
    
    second_row = ["Title", "Predict Precision", "Actual Genre", "Predict #1", "Predict #2", "Predict #3" ]
    print(str(second_row))
    sheet1.append(second_row)
    # sheet1에 list추가.

    result_len = len(result_list)
    result_sum = 0
    result_average = 0
    for result in result_list:
    # Result :
    # [ '타이틀' ' 범위 이내 맞춘 빈도', 실제 장르, '정확도 예측값', '맞춘 것만' ]
        # 정확도 계산에 사용됨. 
        tmp_len = len(result[2]) 
        if tmp_len >=4 :
            tmp_len = 3
        result_row = [result[0], result[1]/tmp_len, str(result[2]), result[3][0][0], result[3][1][0], result[3][2][0] ]
        print(str(result_row))
        sheet1.append( result_row )
        result_sum += result_row[1]
    result_average = result_sum / result_len
    sheet1.append(["average precision :", result_average])

    last_row = ["#########", "#########"]
    print(str(last_row))
    sheet1.append(last_row)
    # excel 저장.
    wb.save(filename=file_name)


def print_and_save_result_to_excel_hl(movie_names, true_labels, pred_labels, hl_results, hl_average):
    file_name = 'result_hammingloss.xlsx'

    if os.path.isfile(file_name):
        # file 있을시 load
        wb = openpyxl.load_workbook(file_name)
        # 데이터를 넣을 시트를 활성화합니다.
        sheet1 = wb.active
    else:
        # file 없을시 생성
        wb = openpyxl.Workbook()
        # 데이터를 넣을 시트를 활성화합니다.
        sheet1 = wb.active
        # 시트의 이름을 정합니다.
        sheet1.title = 'resultSheet'

    first_row = ["#########", "TESTIFY"]
    print(str(first_row))
    sheet1.append(first_row)

    second_row = ["Title", "Hamming Loss", "Actual Genre", "Predict Genre"]
    print(str(second_row))
    sheet1.append(second_row)

    # 각 영화별 hamming loss 값 및 실제/예측 장르 저장
    for (i, movie) in zip(range(len(movie_names)), movie_names):
        result_row = [movie, hl_results[i], str(true_labels[i]), str(sorted(pred_labels[i]))]
        print(str(result_row))
        sheet1.append(result_row)

    average_row = ["average hl", hl_average]
    print(str(average_row))
    sheet1.append(average_row)

    last_row = ["#########", "#########"]
    print(str(last_row))
    sheet1.append(last_row)

    # excel 저장.
    wb.save(filename=file_name)


def print_and_save_result_to_excel_pr(label_names, precisions_per_label, recalls_per_label):
    file_name = 'result_precision_recall.xlsx'

    if os.path.isfile(file_name):
        # file 있을시 load
        wb = openpyxl.load_workbook(file_name)
        # 데이터를 넣을 시트를 활성화합니다.
        sheet1 = wb.active
    else:
        # file 없을시 생성
        wb = openpyxl.Workbook()
        # 데이터를 넣을 시트를 활성화합니다.
        sheet1 = wb.active
        # 시트의 이름을 정합니다.
        sheet1.title = 'resultSheet'

    first_row = ["#########", "TESTIFY"]
    print(str(first_row))
    sheet1.append(first_row)

    second_row = ["Genre", "Precision", "Recall"]
    print(str(second_row))
    sheet1.append(second_row)

    precision_avg = 0
    recall_avg = 0
    actual_label_cnt = 0
    # 각 영화별 hamming loss 값 및 실제/예측 장르 저장
    for (i, label) in zip(range(len(label_names)), label_names):
        if recalls_per_label[i] != -1:
            result_row = [label, precisions_per_label[i], recalls_per_label[i]]
            print(str(result_row))
            sheet1.append(result_row)

            precision_avg += precisions_per_label[i]
            recall_avg += recalls_per_label[i]
            actual_label_cnt += 1

    precision_avg /= actual_label_cnt
    recall_avg /= actual_label_cnt
    average_row = ["Average", precision_avg, recall_avg]
    print(str(average_row))
    sheet1.append(average_row)

    last_row = ["#########", "#########"]
    print(str(last_row))
    sheet1.append(last_row)

    # excel 저장.
    wb.save(filename=file_name)

